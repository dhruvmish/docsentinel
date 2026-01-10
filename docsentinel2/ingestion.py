from dataclasses import dataclass, field
from typing import List, Tuple, Any
import fitz  # PyMuPDF
import os
import re
import openpyxl


# ---------- PDF structured representation ----------
@dataclass
class Block:
    id: str
    type: str  # "text" or "image"
    content: str = ""
    path: str = ""
    bbox: Tuple[float, float, float, float] = (0, 0, 0, 0)
    page_number: int = 0


@dataclass
class PageRep:
    number: int
    blocks: List[Block] = field(default_factory=list)


@dataclass
class DocumentRep:
    path: str
    pages: List[PageRep] = field(default_factory=list)


# ---------- XLSX structured representation ----------
@dataclass
class CellRep:
    row: int
    col: int
    value: Any = None


@dataclass
class TableRep:
    sheet_name: str
    cells: List[CellRep] = field(default_factory=list)
    max_row: int = 0
    max_col: int = 0


@dataclass
class WorkbookRep:
    path: str
    tables: List[TableRep] = field(default_factory=list)


# ---------- Helpers ----------
def _extract_text_from_pdf(path: str) -> str:
    doc = fitz.open(path)
    texts = []
    for page in doc:
        texts.append(page.get_text())
    doc.close()
    return "\n".join(texts)


def _split_into_sentences(text: str) -> List[str]:
    text = text.replace("\n", " ")
    sentences = re.split(r'(?<=[.!?])\s+', text)
    return [s.strip() for s in sentences if s.strip()]


def _trim_bounds(ws):
    min_row, min_col = ws.max_row, ws.max_column
    max_row, max_col = 0, 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in [None, ""]:
                min_row = min(min_row, cell.row)
                min_col = min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)

    return min_row, min_col, max_row, max_col


def load_xlsx(path: str) -> WorkbookRep:
    wb = openpyxl.load_workbook(path, data_only=True)
    tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        min_r, min_c, max_r, max_c = _trim_bounds(ws)
        if max_r == 0:
            continue

        table_cells = []
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                val = ws.cell(row=r, column=c).value
                if val not in [None, ""]:
                    table_cells.append(CellRep(row=r, col=c, value=val))

        tables.append(TableRep(
            sheet_name=sheet_name,
            cells=table_cells,
            max_row=max_r - min_r + 1,
            max_col=max_c - min_c + 1
        ))

    return WorkbookRep(path=path, tables=tables)


# ---------- Sentence + page + bbox for PDFs ----------
def load_pdf_sentences_with_positions(path: str) -> List[dict]:
    doc = fitz.open(path)
    results: List[dict] = []

    for page_idx, page in enumerate(doc):
        blocks = page.get_text("blocks")
        for block in blocks:
            x0, y0, x1, y1, text, *_ = block
            text = text.strip()
            if not text:
                continue

            sentences = _split_into_sentences(text)
            for sent in sentences:
                if not sent:
                    continue
                results.append({
                    "sentence": sent,
                    "page": page_idx + 1,
                    "bbox": (x0, y0, x1, y1),
                })

    doc.close()
    return results


# ---------- Public API ----------
def load_document(path: str):
    ext = os.path.splitext(path)[1].lower()

    if ext == ".pdf":
        sent_meta = load_pdf_sentences_with_positions(path)
        return [m["sentence"] for m in sent_meta]

    elif ext == ".xlsx":
        return load_xlsx(path)

    else:
        raise ValueError(f"Unsupported file type: {ext}")


# ---------- PDF layout (TEXT + IMAGE BLOCKS) ----------
def load_pdf_layout(path: str, image_output_dir: str) -> DocumentRep:
    os.makedirs(image_output_dir, exist_ok=True)
    doc = fitz.open(path)
    pages: List[PageRep] = []

    for page_idx, page in enumerate(doc):
        page_number = page_idx + 1
        page_blocks: List[Block] = []

        # -------- TEXT BLOCKS (UNCHANGED) --------
        for i, block in enumerate(page.get_text("blocks")):
            x0, y0, x1, y1, text, *_ = block
            text = text.strip()
            if not text:
                continue
            block_id = f"p{page_number}_t{i}"
            page_blocks.append(
                Block(
                    id=block_id,
                    type="text",
                    content=text,
                    bbox=(x0, y0, x1, y1),
                    page_number=page_number
                )
            )

        # -------- IMAGE BLOCKS (FIXED) --------
        for j, img in enumerate(page.get_images(full=True)):
            xref = img[0]

            pix = fitz.Pixmap(doc, xref)

            # 🔧 FORCE colorspace to RGB (handles CMYK, ICC, Separation, etc.)
            if pix.colorspace is None or pix.colorspace.name != "DeviceRGB":
                pix = fitz.Pixmap(fitz.csRGB, pix)

            # Remove alpha channel if present
            if pix.alpha:
                pix = fitz.Pixmap(pix, 0)

            img_filename = f"p{page_number}_img{j}.png"
            img_path = os.path.join(image_output_dir, img_filename)

            pix.save(img_path)
            pix = None

            # ✅ CORRECT: extract image rectangle(s) on page
            rects = page.get_image_rects(xref)
            if rects:
                r = rects[0]
                bbox = (r.x0, r.y0, r.x1, r.y1)
            else:
                # Safe fallback (rare)
                pr = page.rect
                bbox = (pr.x0, pr.y0, pr.x1, pr.y1)

            block_id = f"p{page_number}_i{j}"
            page_blocks.append(
                Block(
                    id=block_id,
                    type="image",
                    path=img_path,
                    bbox=bbox,
                    page_number=page_number
                )
            )

        pages.append(PageRep(number=page_number, blocks=page_blocks))

    doc.close()
    return DocumentRep(path=path, pages=pages)
